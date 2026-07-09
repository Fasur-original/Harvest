"""Named data-layer functions for songs (PDD §6.1)."""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from typing import BinaryIO

import openpyxl
from openpyxl.styles import Font
from openpyxl.workbook import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models import Song, SongLine


@dataclass
class ParsedSong:
    title: str
    lines: list[dict]  # [{"line_number": int, "line_text": str, "repeat_count": int}, ...]
    artist: str | None = None


@dataclass
class SheetError:
    tab: str
    problem: str


@dataclass
class ParsedSongSheet:
    songs: list[ParsedSong]
    errors: list[SheetError]


def parse_song_sheet(file: BinaryIO | Workbook) -> ParsedSongSheet:
    """Reads one workbook, one tab per song, tab name = title (PDD §10.2).

    Each tab needs a header row naming its columns. Only "line_text" is
    required; "repeat_count" is optional and defaults to 1 when blank, since
    most lines don't repeat and typing "1" on every single row is exactly
    the kind of tedium worth designing out. There's no "line_number" column
    at all -- line order comes from row order, one less thing for whoever
    fills the sheet in to get wrong or keep in sync if rows get reordered
    later.

    Best-effort (PDD §10.3): a malformed tab is recorded as an error naming
    that exact tab and problem, and skipped -- it doesn't fail every other
    tab in the same workbook.

    Accepts either a raw file or an already-opened Workbook -- the bulk
    import dispatcher (parse_song_import) has to open the workbook itself
    first to sniff whether it's this tab-per-song format or the flat
    format, and shouldn't need to read the same file twice to find out.
    """
    workbook = file if isinstance(file, Workbook) else openpyxl.load_workbook(file, data_only=True, read_only=True)

    songs: list[ParsedSong] = []
    errors: list[SheetError] = []
    seen_titles: set[str] = set()

    for sheet in workbook.worksheets:
        title = sheet.title.strip()

        if title in seen_titles:
            errors.append(SheetError(tab=title, problem=f'duplicate tab name "{title}" -- each song needs a unique tab'))
            continue
        seen_titles.add(title)

        rows = list(sheet.iter_rows(values_only=True))
        if not rows or all(cell is None for cell in rows[0]):
            errors.append(SheetError(tab=title, problem="empty sheet, no header row"))
            continue

        header = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
        if "line_text" not in header:
            errors.append(SheetError(tab=title, problem='missing required "line_text" column in the header row'))
            continue

        text_col = header.index("line_text")
        repeat_col = header.index("repeat_count") if "repeat_count" in header else None

        lines: list[dict] = []
        row_problem: str | None = None
        for row_index, row in enumerate(rows[1:], start=2):
            text = row[text_col] if text_col < len(row) else None
            if text is None or str(text).strip() == "":
                continue  # skip blank rows silently -- common spreadsheet padding

            repeat_count = 1
            if repeat_col is not None and repeat_col < len(row) and row[repeat_col] is not None:
                raw = row[repeat_col]
                try:
                    repeat_count = int(raw)
                    if repeat_count < 1:
                        raise ValueError
                except (TypeError, ValueError):
                    row_problem = f'row {row_index}: "repeat_count" must be a positive whole number, got {raw!r}'
                    break

            lines.append(
                {"line_number": len(lines) + 1, "line_text": str(text).strip(), "repeat_count": repeat_count}
            )

        if row_problem is not None:
            errors.append(SheetError(tab=title, problem=row_problem))
            continue

        if not lines:
            errors.append(SheetError(tab=title, problem="no lyric lines found under the header row"))
            continue

        songs.append(ParsedSong(title=title, lines=lines))

    return ParsedSongSheet(songs=songs, errors=errors)


_FLAT_REQUIRED_COLUMNS = {"title", "lyrics"}


def _flat_rows_to_songs(header: list, rows: list[tuple]) -> ParsedSongSheet:
    """Shared row-walking logic for the flat title/artist/lyrics bulk-import
    format -- one row per song, lyrics as a single multi-line cell -- used
    by both the CSV and single-sheet XLSX variants, since both ultimately
    reduce to the same header + row-tuple shape. Far less tedious to
    prepare for a large batch than the tab-per-song workbook format
    (Phase 07), which is kept as-is and still supported alongside this.
    """
    normalized_header = [str(c).strip().lower() if c is not None else "" for c in header]
    if not _FLAT_REQUIRED_COLUMNS.issubset(normalized_header):
        missing = _FLAT_REQUIRED_COLUMNS - set(normalized_header)
        return ParsedSongSheet(
            songs=[],
            errors=[SheetError(tab="header row", problem=f"missing required column(s): {', '.join(sorted(missing))}")],
        )

    title_col = normalized_header.index("title")
    artist_col = normalized_header.index("artist") if "artist" in normalized_header else None
    lyrics_col = normalized_header.index("lyrics")

    songs: list[ParsedSong] = []
    errors: list[SheetError] = []
    seen_titles: set[str] = set()

    for row_index, row in enumerate(rows, start=2):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue  # skip blank rows silently, same as the tab-per-song path

        title = str(row[title_col]).strip() if title_col < len(row) and row[title_col] is not None else ""
        label = title or f"row {row_index}"

        if not title:
            errors.append(SheetError(tab=label, problem=f"row {row_index}: missing a song title"))
            continue

        if title in seen_titles:
            errors.append(SheetError(tab=label, problem=f'duplicate title "{title}" in this import'))
            continue
        seen_titles.add(title)

        artist = None
        if artist_col is not None and artist_col < len(row) and row[artist_col] is not None:
            artist = str(row[artist_col]).strip() or None

        raw_lyrics = row[lyrics_col] if lyrics_col < len(row) else None
        if raw_lyrics is None or str(raw_lyrics).strip() == "":
            errors.append(SheetError(tab=label, problem=f'"{title}": missing lyrics'))
            continue

        lines = [
            {"line_number": i + 1, "line_text": line.strip(), "repeat_count": 1}
            for i, line in enumerate(str(raw_lyrics).splitlines())
            if line.strip()
        ]
        if not lines:
            errors.append(SheetError(tab=label, problem=f'"{title}": lyrics cell has no non-blank lines'))
            continue

        songs.append(ParsedSong(title=title, artist=artist, lines=lines))

    return ParsedSongSheet(songs=songs, errors=errors)


def parse_song_csv(file: BinaryIO) -> ParsedSongSheet:
    """Flat bulk-import format: one row per song, columns title/artist/lyrics
    (artist optional), lyrics as a single cell with each lyric line on its
    own line inside the cell.
    """
    text = file.read().decode("utf-8-sig")  # -sig strips a BOM if Excel/Numbers added one
    rows = list(csv.reader(io.StringIO(text)))
    if not rows:
        return ParsedSongSheet(songs=[], errors=[SheetError(tab="file", problem="empty file, no header row")])
    return _flat_rows_to_songs(rows[0], rows[1:])


def _is_flat_workbook(workbook: Workbook) -> bool:
    if len(workbook.worksheets) != 1:
        return False
    header_row = next(workbook.worksheets[0].iter_rows(values_only=True, max_row=1), None)
    if header_row is None:
        return False
    normalized = {str(c).strip().lower() for c in header_row if c is not None}
    return _FLAT_REQUIRED_COLUMNS.issubset(normalized)


def _parse_flat_xlsx(workbook: Workbook) -> ParsedSongSheet:
    rows = list(workbook.worksheets[0].iter_rows(values_only=True))
    return _flat_rows_to_songs(rows[0], rows[1:])


def parse_song_import(file: BinaryIO, filename: str) -> ParsedSongSheet:
    """Dispatches a bulk-import upload to the right parser. A `.csv` is
    always the flat title/artist/lyrics format; a `.xlsx` is sniffed by its
    header row, since it can be either that same flat format (a single
    sheet) or the original tab-per-song workbook format (Phase 07) -- both
    are real, still-supported inputs, not a replacement of one by the other.
    """
    if filename.lower().endswith(".csv"):
        return parse_song_csv(file)

    workbook = openpyxl.load_workbook(file, data_only=True, read_only=True)
    if _is_flat_workbook(workbook):
        return _parse_flat_xlsx(workbook)
    return parse_song_sheet(workbook)


def build_flat_import_template() -> bytes:
    """A ready-to-edit example CSV for the flat title/artist/lyrics format --
    the counterpart to build_song_sheet_template() below for the original
    tab-per-song workbook format.
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["title", "artist", "lyrics"])
    writer.writerow(
        [
            "Amazing Grace",
            "John Newton",
            "Amazing grace, how sweet the sound\n"
            "That saved a wretch like me\n"
            "I once was lost, but now am found\n"
            "Was blind, but now I see",
        ]
    )
    return buffer.getvalue().encode("utf-8")


def build_song_sheet_template() -> bytes:
    """A ready-to-duplicate example workbook (PDD §10.3).

    Whoever preps the sheet copies this tab for each new song and renames it
    (Excel: right-click -> Move or Copy -> Create a copy) rather than
    building the column layout from scratch every time. Shows both an
    ordinary line and a repeated one, so the `repeat_count` column's meaning
    is obvious from the example rather than needing separate instructions.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Example Song (copy this tab)"

    sheet.append(["line_text", "repeat_count"])
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    sheet.append(["Amazing grace, how sweet the sound", None])
    sheet.append(["That saved a wretch like me", None])
    sheet.append(["I once was lost, but now am found", 3])
    sheet.append(["Was blind, but now I see", None])

    sheet.column_dimensions["A"].width = 42
    sheet.column_dimensions["B"].width = 14

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


async def save_song(db: AsyncSession, title: str, lines: list[dict], artist: str | None = None) -> Song:
    """Creates a song, or replaces an existing one with the same title.

    Re-uploading a song already in the permanent library (PDD §10.5 -- a
    repeat song shouldn't need a tab prepared again, but nothing stops
    someone from including it anyway) updates its lines in place instead of
    creating a duplicate entry that would clutter search results.
    """
    existing = (
        await db.execute(select(Song).where(Song.title == title).options(selectinload(Song.lines)))
    ).scalar_one_or_none()

    new_lines = [
        SongLine(
            line_number=line["line_number"],
            line_text=line["line_text"],
            repeat_count=line.get("repeat_count", 1),
        )
        for line in lines
    ]

    if existing is not None:
        existing.lines = new_lines  # cascade="all, delete-orphan" replaces the old rows
        # Only overwrite artist if this import actually named one -- a
        # tab-per-song re-upload (which never carries an artist) shouldn't
        # silently wipe metadata a flat import or manual edit set earlier.
        if artist is not None:
            existing.artist = artist
        song = existing
    else:
        song = Song(title=title, artist=artist, lines=new_lines)
        db.add(song)

    await db.commit()
    await db.refresh(song, attribute_names=["lines"])
    return song


async def get_song(db: AsyncSession, song_id: int) -> Song | None:
    result = await db.execute(
        select(Song).where(Song.id == song_id).options(selectinload(Song.lines))
    )
    return result.scalar_one_or_none()


async def search_songs(db: AsyncSession, query: str) -> list[Song]:
    result = await db.execute(select(Song).where(Song.title.ilike(f"%{query}%")))
    return list(result.scalars())
